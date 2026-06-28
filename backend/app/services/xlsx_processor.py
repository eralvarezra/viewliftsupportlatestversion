# backend/app/services/xlsx_processor.py
"""Excel file processor for FAQ and reference knowledge base."""
import os
from typing import List, Dict, Any
from openpyxl import load_workbook


class XlsxProcessor:
    """Process XLSX files and extract text chunks from tabular data.

    Supports:
    - FAQ format: Question/Answer columns
    - Reference data: Any tabular data (zip codes, service areas, etc.)

    Each row is converted to a text chunk for embedding.
    """

    # Token estimation: approximately 4 characters per token for English
    CHARS_PER_TOKEN = 4

    def __init__(self, max_tokens: int = 500):
        if max_tokens <= 0:
            raise ValueError("max_tokens must be a positive integer")
        self.max_tokens = max_tokens

    def extract_rows(self, filepath: str) -> List[Dict[str, str]]:
        """Extract all rows from an XLSX file as dictionaries.

        Args:
            filepath: Path to the XLSX file.

        Returns:
            List of dictionaries, one per row, with header names as keys.
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"File not found: {filepath}")

        entries = []
        workbook = None
        try:
            workbook = load_workbook(filepath, read_only=True, data_only=True)
            sheet = workbook.active
            if not sheet:
                raise ValueError(f"No active sheet in: {filepath}")

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(f"Column_{cell.column}")

            # Extract data rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue

                entry = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        key = headers[idx]
                        entry[key] = str(value).strip() if value else ""

                # Only include entries with at least one non-empty value
                if any(entry.values()):
                    entries.append(entry)
        except Exception as e:
            raise ValueError(f"Invalid XLSX file: {filepath}") from e
        finally:
            if workbook:
                workbook.close()

        return entries

    def row_to_text(self, row: Dict[str, str]) -> str:
        """Convert a row dictionary to text for embedding.

        Handles both FAQ format and generic reference data.

        Args:
            row: Dictionary with row data.

        Returns:
            Formatted text string.
        """
        # Check if this looks like FAQ format (has Question/Answer or Spanish equivalents)
        headers_lower = {k.lower(): k for k in row.keys()}
        is_faq = any(h in headers_lower for h in ['question', 'pregunta', 'answer', 'respuesta'])

        if is_faq:
            # FAQ format - prioritize Q&A
            question = row.get(headers_lower.get('question', '')) or row.get(headers_lower.get('pregunta', '')) or ""
            answer = row.get(headers_lower.get('answer', '')) or row.get(headers_lower.get('respuesta', '')) or ""

            parts = []
            if question:
                parts.append(f"Question: {question}")
            if answer:
                parts.append(f"Answer: {answer}")

            # Include other columns as context
            for key, value in row.items():
                if key.lower() not in ['question', 'pregunta', 'answer', 'respuesta'] and value:
                    parts.append(f"{key}: {value}")

            return "\n".join(parts)
        else:
            # Reference data format - convert entire row to searchable text
            parts = []
            for key, value in row.items():
                if value:
                    parts.append(f"{key}: {value}")
            return " | ".join(parts)

    def chunk_text(self, text: str, max_tokens: int = None) -> List[str]:
        """Split text into chunks of approximately max_tokens.

        Args:
            text: Text to chunk.
            max_tokens: Maximum tokens per chunk.

        Returns:
            List of text chunks.
        """
        if max_tokens is None:
            max_tokens = self.max_tokens
        if max_tokens <= 0:
            raise ValueError("max_tokens must be a positive integer")

        max_chars = max_tokens * self.CHARS_PER_TOKEN

        words = text.split()
        chunks = []
        current_chunk = []
        current_length = 0

        for word in words:
            word_length = len(word) + 1  # +1 for space
            if current_length + word_length > max_chars and current_chunk:
                chunks.append(" ".join(current_chunk))
                current_chunk = [word]
                current_length = word_length
            else:
                current_chunk.append(word)
                current_length += word_length

        if current_chunk:
            chunks.append(" ".join(current_chunk))

        return chunks

    def process_xlsx(self, filepath: str) -> Dict[str, Any]:
        """Process an XLSX file and return chunks.

        Args:
            filepath: Path to the XLSX file.

        Returns:
            Dictionary with:
            - filename: Base filename
            - chunks: List of text chunks
            - total_chunks: Number of chunks
            - entries_count: Number of rows extracted
        """
        rows = self.extract_rows(filepath)
        chunks = []

        for row in rows:
            text = self.row_to_text(row)
            if text.strip():
                # Each row is its own chunk (they're typically small)
                # But if a row is too large, split it
                row_chunks = self.chunk_text(text)
                chunks.extend(row_chunks)

        return {
            "filename": os.path.basename(filepath),
            "chunks": chunks,
            "total_chunks": len(chunks),
            "entries_count": len(rows)
        }