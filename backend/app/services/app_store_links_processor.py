"""Dedicated processor for the All Clients App Details - App Store Links spreadsheet."""
from typing import Dict, Any, List
from openpyxl import load_workbook

PLATFORM_COLS = {
    1:  "Website",
    2:  "iOS App Store",
    3:  "Apple TV App Store",
    4:  "Android (Google Play)",
    5:  "Roku Channel Store",
    6:  "Amazon Fire TV",
    7:  "Fire TV VegaOS",
    8:  "Android TV / Google TV",
    9:  "Samsung Smart TV",
    10: "LG Smart TV",
    11: "Vizio Smart TV",
    12: "Xbox",
    13: "Xfinity & Xumo TV",
    15: "One Trust",
    16: "Internal Web Consent",
}

SKIP_VALUES = {
    "na", "n/a", "none", "not live", "(not live)", "coming soon!", "coming soon",
    "not launched", "not this season", "legacy old app", "informed", "pending",
    "x", "", "nan",
}


def _is_skip(val: str) -> bool:
    return val.strip().lower().strip("()") in SKIP_VALUES or not val.strip()


def _clean(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    # Remove newlines used as separators in cells, keep first URL if multiple
    lines = [l.strip() for l in s.splitlines() if l.strip()]
    return lines[0] if lines else ""


def _extra_notes(val) -> str:
    """Extract notes embedded after newlines in a cell (e.g. IAP = True)."""
    if val is None:
        return ""
    lines = [l.strip() for l in str(val).splitlines() if l.strip()]
    return " | ".join(lines[1:]) if len(lines) > 1 else ""


class AppStoreLinksProcessor:
    """Process the All Clients App Details spreadsheet into per-client chunks."""

    @staticmethod
    def detect(filepath: str) -> bool:
        try:
            wb = load_workbook(filepath, read_only=True)
            try:
                return "App Store Links" in wb.sheetnames
            finally:
                wb.close()
        except Exception:
            return False

    def process(self, filepath: str) -> Dict[str, Any]:
        chunks = []
        chunks.extend(self._process_app_store_sheet(filepath))
        chunks.extend(self._process_samsung_vizio_lg_sheet(filepath))
        return {
            "filename": filepath,
            "chunks": chunks,
            "total_chunks": len(chunks),
            "entries_count": len(chunks),
        }

    def _process_app_store_sheet(self, filepath: str) -> List[str]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        try:
            ws = wb["App Store Links"]
            chunks = []
            current_section = "CLIENT OWNED"

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                # Row 0 (index): section label row
                if row_idx == 0:
                    if row[0]:
                        current_section = str(row[0]).strip()
                    continue
                # Row 1: column headers — skip
                if row_idx == 1:
                    continue

                client = _clean(row[0])
                if not client or client.lower() in SKIP_VALUES:
                    # Could be a new section header
                    if row[0] and not any(row[1:]):
                        current_section = str(row[0]).strip()
                    continue

                lines = [f"Client: {client}", f"Category: {current_section}"]

                for col_idx, label in PLATFORM_COLS.items():
                    if col_idx >= len(row):
                        continue
                    raw = row[col_idx]
                    val = _clean(raw)
                    notes = _extra_notes(raw)
                    if _is_skip(val):
                        continue
                    entry = f"{label}: {val}"
                    if notes and not _is_skip(notes):
                        entry += f" ({notes})"
                    lines.append(entry)

                # Only create a chunk if there's at least one link besides client/category
                if len(lines) > 2:
                    chunks.append("\n".join(lines))

            return chunks
        finally:
            wb.close()

    def _process_samsung_vizio_lg_sheet(self, filepath: str) -> List[str]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        try:
            if "SamsungVizioLG" not in wb.sheetnames:
                return []
            ws = wb["SamsungVizioLG"]
            chunks = []

            SMART_TV_COLS = {
                1: "Samsung Smart TV install URL",
                2: "Vizio Smart TV install URL",
                3: "LG Smart TV install URL",
                4: "Fire TV VegaOS install URL",
            }

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    continue  # headers
                client = _clean(row[0])
                if not client or client.lower() in SKIP_VALUES:
                    continue

                lines = [
                    f"Client: {client}",
                    "[INTERNAL - Smart TV install URLs, do not share with customers]",
                ]

                for col_idx, label in SMART_TV_COLS.items():
                    if col_idx >= len(row):
                        continue
                    raw = row[col_idx]
                    val = _clean(raw)
                    notes = _extra_notes(raw)
                    if _is_skip(val):
                        continue
                    # Extract model year note from Samsung column
                    entry = f"{label}: {val}"
                    if notes and not _is_skip(notes):
                        entry += f" ({notes})"
                    lines.append(entry)

                if len(lines) > 2:
                    chunks.append("\n".join(lines))

            return chunks
        finally:
            wb.close()
